#!/usr/bin/env python3
import os
import re
from openpyxl import Workbook, load_workbook


BASE_DIR = os.path.join(os.path.dirname(__file__), "Примеры ручного анализа")
OUT_FILE = os.path.join(os.path.dirname(__file__), "Config_по_электростанциям.xlsx")


def normalize(text: str) -> str:
    s = (text or "").strip().upper().replace("Ё", "Е")
    s = re.sub(r"\s+", " ", s)
    s = s.replace("Э", "")
    return s


def strip_power_prefix(header: str) -> str:
    t = (header or "").strip()
    t = re.sub(r"^[РP]т\s*_?", "", t, flags=re.IGNORECASE)
    t = re.sub(r"^[РP]\s*_?", "", t, flags=re.IGNORECASE)
    return t.replace("_", " ").strip()


def strip_freq_prefix(header: str) -> str:
    t = (header or "").strip()
    t = re.sub(r"^[fF]\s*_?", "", t)
    return t.replace("_", " ").strip()


def station_from_filename(filename: str) -> str:
    name = filename.replace("Осциллограммы - ", "")
    name = re.sub(r"\s+\d{2}\.\d{2}\.\d{4}\.xlsx$", "", name)
    return name.strip()


def collect_records(base_dir: str):
    files = [
        f
        for f in os.listdir(base_dir)
        if f.lower().endswith(".xlsx") and f.startswith("Осциллограммы - ")
    ]

    records = []

    for filename in sorted(files):
        station = station_from_filename(filename)
        path = os.path.join(base_dir, filename)
        wb = load_workbook(path, data_only=True, read_only=True)
        if "ОИК" not in wb.sheetnames:
            continue

        ws = wb["ОИК"]
        top = [[ws.cell(r, c).value for c in range(1, 60)] for r in range(1, 7)]
        header = top[2]  # row 3

        freq_cols = []
        for idx, val in enumerate(header, 1):
            s = str(val).strip() if val is not None else ""
            if s.lower().startswith("f"):
                freq_cols.append((idx, s))

        gen_cols = []
        for idx, val in enumerate(header, 1):
            s = str(val).strip() if val is not None else ""
            if not re.match(r"^[РP]\s*[_ ]", s):
                continue

            low = s.lower().replace("_", " ")
            if re.match(r"^[рp]т\b", low):
                continue
            if any(
                token in low
                for token in [" рэнерг", " рутилиз", " р пгрэс", " р утэц", " р ятэц", " р слпк"]
            ):
                continue
            if ("тг" in low) or ("блок" in low):
                gen_cols.append((idx, s))

        # Read S and Pном from the Рт/Pт area in OIK (if present).
        rt_map = {}
        current_s = None
        current_pnom = None
        for idx, val in enumerate(header, 1):
            s = str(val).strip() if val is not None else ""
            if not re.match(r"^[РP]т\s*", s, re.IGNORECASE):
                continue

            left_r4 = top[3][idx - 2] if idx - 2 >= 0 else None
            left_r5 = top[4][idx - 2] if idx - 2 >= 0 else None

            if isinstance(left_r4, str) and left_r4.replace(" ", "").upper().startswith("S,"):
                current_s = top[3][idx - 1]
            elif isinstance(top[3][idx - 1], (int, float)) and current_s is not None:
                current_s = top[3][idx - 1]

            if isinstance(left_r5, str) and "PНОМ" in left_r5.replace(" ", "").upper():
                current_pnom = top[4][idx - 1]
            elif isinstance(top[4][idx - 1], (int, float)) and current_pnom is not None:
                current_pnom = top[4][idx - 1]

            rt_key = normalize(strip_power_prefix(s))
            rt_map[rt_key] = {"S": current_s, "Pном": current_pnom}

        default_freq = freq_cols[0][1] if freq_cols else "Частота"

        for _, power_header in gen_cols:
            generator = strip_power_prefix(power_header)
            nkey = normalize(generator)

            freq_header = default_freq
            for _, freq_h in freq_cols:
                if normalize(strip_freq_prefix(freq_h)) == nkey:
                    freq_header = freq_h
                    break

            vals = rt_map.get(nkey, {})
            records.append(
                {
                    "Станция": station,
                    "Генератор": generator,
                    "Колонка_мощности": power_header,
                    "Колонка_частоты": freq_header,
                    "Тип_оборудования": "ПТУ",
                    "Pном, МВт": vals.get("Pном", ""),
                    "S, %": vals.get("S", ""),
                    "fнч, Гц": "",
                    "Kд": 0.5,
                    "Вкл (1/0)": 1,
                    "Кач_вкл (1/0)": 1,
                    "t5, c": 15,
                    "dP5, %Pном": 5,
                    "t10, c": 420,
                    "dP10, %Pном": 10,
                    "Уст_допуск, %Pном": 1,
                    "В сумму станции (1/0)": 1 if any(k in station.upper() for k in ["СОСНОГОР", "ВОРКУТ", "СЛПК"]) else 0,
                    "Источник_параметров": "oik_embedded" if vals else "",
                    "Файл": filename,
                }
            )

    return records


def write_output(path: str, records):
    headers = [
        "Станция",
        "Генератор",
        "Колонка_мощности",
        "Колонка_частоты",
        "Тип_оборудования",
        "Pном, МВт",
        "S, %",
        "fнч, Гц",
        "Kд",
        "Вкл (1/0)",
        "Кач_вкл (1/0)",
        "t5, c",
        "dP5, %Pном",
        "t10, c",
        "dP10, %Pном",
        "Уст_допуск, %Pном",
        "В сумму станции (1/0)",
        "Источник_параметров",
        "Файл",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "AllStations"
    ws.append(headers)
    for rec in records:
        ws.append([rec[h] for h in headers])

    # Sheet ready to copy directly into Excel macro Config (A:Q + S:T settings).
    cfg = wb.create_sheet("Config_For_Macro")
    cfg.append(headers[:-2])  # without helper metadata columns
    for rec in records:
        cfg.append([rec[h] for h in headers[:-2]])
    cfg["S1"] = "Глобальные настройки"
    cfg["S2"] = "fном, Гц"
    cfg["T2"] = 50
    cfg["S3"] = "Время начала события"
    cfg["T3"] = ""
    cfg["S4"] = "Автопоиск старта (1/0)"
    cfg["T4"] = 1
    cfg["S5"] = "Колич. интервал, с"
    cfg["T5"] = 82
    cfg["S6"] = "Допуск количеств., %"
    cfg["T6"] = 10
    cfg["S7"] = "Порог включения в работу, МВт"
    cfg["T7"] = 1

    grouped = {}
    for rec in records:
        grouped.setdefault(rec["Станция"], []).append(rec)

    for station, rows in grouped.items():
        sname = re.sub(r"[:\\/?*\[\]]", "_", station)[:31]
        w = wb.create_sheet(sname)
        w.append(headers)
        for rec in rows:
            w.append([rec[h] for h in headers])

    for w in wb.worksheets:
        for col in w.columns:
            letter = col[0].column_letter
            width = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            w.column_dimensions[letter].width = min(max(width + 2, 10), 60)

    wb.save(path)


def main():
    records = collect_records(BASE_DIR)
    write_output(OUT_FILE, records)

    by_station = {}
    for rec in records:
        by_station.setdefault(rec["Станция"], []).append(rec)

    print("Saved:", OUT_FILE)
    print("Rows:", len(records))
    for station in sorted(by_station):
        rows = by_station[station]
        filled = sum(1 for r in rows if r["Pном, МВт"] not in ("", None) and r["S, %"] not in ("", None))
        print(f"- {station}: {len(rows)} generators, filled Pном+S = {filled}")


if __name__ == "__main__":
    main()
